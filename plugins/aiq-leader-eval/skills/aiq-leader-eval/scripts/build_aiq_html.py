#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""AI-Q 리더 평가 가이드 v7 HTML 빌더 (경영관리 v7 양식 기반)."""
import argparse
import csv
import html
import io
import json
import os
import re
import sys

try:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
except Exception:
    pass


# ---------- 컬럼 매핑 ----------
COLUMN_KEYWORDS = {
    'email':   ['이메일', 'email'],
    'name':    ['이름'],
    'team':    ['소속', '팀'],
    'role':    ['직책'],
    'leader':  ['리더'],
    'q1_1':    ['가장 많은 시간', '업무 3가지'],
    'q1_2':    ['작업 흐름에 AI'],
    'q1_3':    ['AI가 함께 작동', '비중을 체감'],
    'q2_1':    ['정기적으로', '주 2회'],
    'q2_2':    ['손에 익은'],
    'q2_3':    ['새로 시도해 본'],
    'q3_1':    ['잘 풀렸다', '1순위'],
    'q3_2':    ['1순위 사례의 산출물'],
    'q3_3':    ['2순위'],
    'q3_4':    ['2순위 사례의 산출물'],
    'q3_5':    ['3순위'],
    'q3_6':    ['3순위 사례의 산출물'],
    'q4_1':    ['보통 어떤 흐름', '검토하시나요'],
    'q4_2':    ['오류', '편향'],
    'q4_3':    ['체크리스트', '템플릿'],
    'q5_1':    ['반복 업무가 몇 개', '몇 개 정도'],
    'q5_2':    ['절약되는 시간'],
    'q5_3':    ['엄두를 못', '새로 시도해 본 업무'],
    'q5_4':    ['없었으면', '못 했을'],
    'q5_5':    ['위 사례의 산출물'],
    'q6_1':    ['공유한 경험'],
    'q6_2':    ['옆자리 동료'],
    'q6_3':    ['공유했던 자료'],
    'q8_3':    ['답답하거나', '막막한'],
    'q8_4':    ['가장 큰 어려움'],
    'q8_5':    ['지원이 있으면'],
    'q8_6':    ['자신 있는'],
    'q8_7':    ['다음 3개월'],
}

MAPPING_ORDER = [
    'email', 'name', 'team', 'role', 'leader',
    'q1_1', 'q1_2', 'q1_3',
    'q2_1', 'q2_2', 'q2_3',
    'q3_2', 'q3_4', 'q3_6',
    'q3_1', 'q3_3', 'q3_5',
    'q4_1', 'q4_2', 'q4_3',
    'q5_2', 'q5_5',
    'q5_1', 'q5_3', 'q5_4',
    'q6_1', 'q6_2', 'q6_3',
    'q8_3', 'q8_4', 'q8_5', 'q8_6', 'q8_7',
]


def map_columns(headers):
    used = set()
    mapping = {}
    for key in MAPPING_ORDER:
        keywords = COLUMN_KEYWORDS[key]
        for i, h in enumerate(headers):
            if i in used or not h:
                continue
            h_str = str(h)
            if any(kw in h_str for kw in keywords):
                mapping[key] = i
                used.add(i)
                break
    return mapping


def read_rows(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in ('.xlsx', '.xlsm'):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        return [[c for c in row] for row in ws.iter_rows(values_only=True)]
    elif ext in ('.csv', '.tsv'):
        delim = '\t' if ext == '.tsv' else ','
        with open(path, 'r', encoding='utf-8-sig', newline='') as f:
            return [row for row in csv.reader(f, delimiter=delim)]
    raise ValueError(f'지원하지 않는 파일 형식: {ext}')


# ---------- 신호 분류 ----------
def parse_multi(s):
    if not s:
        return []
    return [p.strip() for p in re.split(r'[,;\n]', s) if p.strip()]


def sig_tool(m):
    n = len(parse_multi(m['q2_1']))
    if n >= 3: return ('green',  '폭넓음', f'{n}종')
    if n == 2: return ('yellow', '복수',   f'{n}종')
    if n == 1: return ('red',    '단일',   f'{n}종')
    return ('red', '없음', '응답 없음')


def sig_freq(m):
    v = m['q1_3']
    if '거의 항상' in v or '75' in v: return ('green',  '체화', v)
    if '자주'      in v or '50' in v: return ('green',  '정착', v)
    if '꾸준히'    in v or '25' in v: return ('yellow', '진행', v)
    if '가끔'      in v or '10' in v: return ('red',    '간헐', v)
    return ('red', '간헐', v or '응답 없음')


def sig_case(m):
    cases = [c for c in [m['q3_1'], m['q3_3'], m['q3_5']] if c]
    n = len(cases)
    avg_len = (sum(len(c) for c in cases) // n) if n else 0
    if n >= 3 and avg_len >= 60: return ('green',  '풍부', f'{n}건 / 평균 {avg_len}자')
    if n >= 2 or (n == 1 and avg_len >= 80): return ('yellow', '보유', f'{n}건 / 평균 {avg_len}자')
    if n >= 1: return ('yellow', '단편', f'{n}건 / 평균 {avg_len}자')
    return ('red', '부재', '응답 없음')


def sig_review(m):
    flow, fixes, checklist = m['q4_1'], m['q4_2'], m['q4_3']
    has_checklist = bool(checklist) and checklist not in [
        '없음', '아직 없음', 'X', 'x', '없습니다', '없어요', '-', '없다'
    ]
    has_fix  = bool(fixes) and len(fixes) > 8
    has_flow = bool(flow)  and len(flow)  > 8
    if has_checklist and (has_fix or has_flow): return ('green',  '체계', '템플릿+검증 흐름')
    if has_checklist:                            return ('yellow', '관리', '체크리스트 보유')
    if has_fix and has_flow:                     return ('yellow', '관리', '검토 흐름·수정 사례')
    if has_flow or has_fix:                      return ('yellow', '의식', '검토 흐름 인식')
    return ('red', '부재', '검토 흐름 약함')


def sig_impact(m):
    auto, time_v, no_ai = m['q5_1'], m['q5_2'], m['q5_4']
    auto_3plus = '3' in auto or '6' in auto
    time_3plus = '3' in time_v or '6' in time_v or '10' in time_v
    has_no_ai  = bool(no_ai) and len(no_ai) > 15
    if auto_3plus and time_3plus and has_no_ai:           return ('green',  '뚜렷', f'{time_v} / {auto}')
    if (auto_3plus and time_3plus) or has_no_ai:           return ('yellow', '실현', f'{time_v} / {auto}')
    if time_v and auto and ('1' in auto or '1' in time_v): return ('yellow', '실현', f'{time_v} / {auto}')
    return ('red', '미미', f'{time_v or "—"} / {auto or "—"}')


def sig_share(m):
    ch, change, link = m['q6_1'], m['q6_2'], m['q6_3']
    channels = parse_multi(ch)
    has_no_share = any('공유한 적 없음' in c or '없음' == c for c in channels)
    n = len([c for c in channels if c and '공유한 적' not in c])
    has_change = bool(change) and len(change) > 10
    has_link   = bool(link)   and 'http' in link.lower()
    if n >= 2 and (has_change or has_link): return ('green',  '확산',   f'{n}가지 방식')
    if n >= 2:                              return ('yellow', '시도',   f'{n}가지 방식')
    if n == 1:                              return ('yellow', '시도',   f'{n}가지 방식')
    if has_no_share:                        return ('red',    '미공유', '공유 활동 없음')
    return ('red', '미공유', '응답 없음')


def normalize_tool(t):
    t = (t or '').strip()
    low = t.lower()
    if 'claude' in low or '클로드' in t: return 'Claude'
    if 'chatgpt' in low or 'gpt' in low: return 'ChatGPT'
    if 'gemini' in low or '제미나이' in t or '제미니' in t: return 'Gemini'
    if 'notion ai' in low or 'notion' in low: return 'Notion AI'
    if 'perplexity' in low or '퍼플렉시티' in t: return 'Perplexity'
    if 'cursor' in low or '커서' in t: return 'Cursor'
    if 'copilot' in low: return 'Copilot'
    if 'midjourney' in low or '미드저니' in t: return 'Midjourney'
    if 'firefly' in low: return 'Firefly'
    if 'n8n' in low: return 'n8n'
    if 'genspark' in low: return 'Genspark'
    if 'figma' in low: return 'Figma AI'
    if 'grok' in low: return 'Grok'
    if '힉스필드' in t or 'hicks' in low or 'higgs' in low: return 'Higgsfield'
    if 'lumi' in low or '루미' in t: return 'Lumi'
    if '사내' in t or '자체' in t: return '사내 자체 툴'
    return t


ROLE_ORDER = {
    '팀장': 0, '팀리더': 1, '파트장': 1, '파트리더': 2, '매니저': 3,
    'MD': 4, 'AMD': 5, 'PM': 6, '시니어': 7, '사원': 8, '팀원': 9, '인턴': 10,
}


def role_sort_key(m):
    role = (m['role'] or '').strip()
    rn = role.upper()
    for k, v in ROLE_ORDER.items():
        if k in role or k.upper() in rn:
            return v
    return 99


SIG_CLASS = {'green': 'sig-green', 'yellow': 'sig-yellow', 'red': 'sig-red'}
SIG_EMOJI = {'green': '🟢', 'yellow': '🟡', 'red': '🔴'}


def esc(s):
    return html.escape('' if s is None else str(s))


def evid(label, val):
    if val and str(val).strip():
        return f'<div class="evidence-item"><div class="q">{esc(label)}</div><div class="a">{esc(val)}</div></div>'
    return f'<div class="evidence-item"><div class="q">{esc(label)}</div><div class="a" style="color:#aaa; font-style:italic;">응답 없음</div></div>'


def score_row(i, sub):
    return (
        f'<div class="score-row">'
        f'<span class="score-prompt">🎯 리더 점수 (1~5):</span>'
        f'<div class="score-buttons" data-member="{i}" data-sub="{sub}">'
        f'<button onclick="setScore({i},\'{sub}\',1)">1</button>'
        f'<button onclick="setScore({i},\'{sub}\',2)">2</button>'
        f'<button onclick="setScore({i},\'{sub}\',3)">3</button>'
        f'<button onclick="setScore({i},\'{sub}\',4)">4</button>'
        f'<button onclick="setScore({i},\'{sub}\',5)">5</button>'
        f'<button class="clear-btn" onclick="setScore({i},\'{sub}\',null)">지우기</button>'
        f'</div></div>'
    )


def signal_line(ax, sub=None):
    s = f"{SIG_EMOJI[ax[0]]} <strong>{esc(ax[1])}</strong> — {esc(ax[2])}"
    if sub is not None:
        s += f" · 보조 신호 {SIG_EMOJI[sub[0]]} {esc(sub[1])}"
    return s


HINT_S22 = '💡 이전 작업물을 직접 비교하거나, 현재 제출물을 BaseLine으로 잡고 다음 분기 평가에 반영'
HINT_S23 = '💡 부서 맥락·이전 사례 반영 작업 → 3점 / 단순 프롬프트만 사용 → 2점'
HINT_S31 = '💡 추가 추적 필요: 어떻게 자동화? (Claude / GPTs / Claude Code / Cowork / n8n 등)'
HINT_S44 = '💡 4-1~3 문항·채널·전파 사례를 종합해서 리더가 판정. 개인 활용에 그치면 2점 / 팀 차원 관행 정착에 기여하면 4~5점'


def hint_box(text):
    return f'<div style="margin:6px 0 10px 0; padding:8px 10px; background:#eef4ff; border-left:3px solid #4a7fdb; font-size:0.88em; color:#2c4670; line-height:1.55;">{esc(text)}</div>'


def build(input_path, dept, output_path):
    rows = read_rows(input_path)
    if not rows or len(rows) < 2:
        raise ValueError('헤더 + 응답 1행 이상이 필요합니다.')

    headers = rows[0]
    col_map = map_columns(headers)

    missing = [k for k in COLUMN_KEYWORDS if k not in col_map]
    if missing:
        print(f'  ⚠️  매핑 실패 컬럼 ({len(missing)}): {missing}', file=sys.stderr)

    def cell(row, key):
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return ''
        v = row[idx]
        return '' if v is None else str(v).strip()

    members = []
    for row in rows[1:]:
        if not any(c is not None and str(c).strip() for c in row):
            continue
        name = cell(row, 'name')
        if not name:
            continue
        m = {'name': name}
        for k in ['role', 'team', 'leader', 'email',
                  'q1_1', 'q1_2', 'q1_3',
                  'q2_1', 'q2_2', 'q2_3',
                  'q3_1', 'q3_2', 'q3_3', 'q3_4', 'q3_5', 'q3_6',
                  'q4_1', 'q4_2', 'q4_3',
                  'q5_1', 'q5_2', 'q5_3', 'q5_4', 'q5_5',
                  'q6_1', 'q6_2', 'q6_3',
                  'q8_3', 'q8_4', 'q8_5', 'q8_6', 'q8_7']:
            m[k] = cell(row, k)
        members.append(m)

    members.sort(key=role_sort_key)
    n_total = len(members)
    print(f'[{dept}] 응답자 {n_total}명')

    sigs = [{
        'tool':   sig_tool(m),
        'freq':   sig_freq(m),
        'case':   sig_case(m),
        'review': sig_review(m),
        'impact': sig_impact(m),
        'share':  sig_share(m),
    } for m in members]

    tool_usage = {}
    for m in members:
        for t in parse_multi(m['q2_1']):
            nt = normalize_tool(t)
            if nt:
                tool_usage[nt] = tool_usage.get(nt, 0) + 1

    freq_dist, share_dist, time_saved_dist, auto_count_dist, difficulty_top = {}, {}, {}, {}, {}
    for m in members:
        v = m['q1_3'] or '응답 없음'; freq_dist[v] = freq_dist.get(v, 0) + 1
        for c in parse_multi(m['q6_1']):
            share_dist[c] = share_dist.get(c, 0) + 1
        v = m['q5_2'] or '응답 없음'; time_saved_dist[v] = time_saved_dist.get(v, 0) + 1
        v = m['q5_1'] or '응답 없음'; auto_count_dist[v] = auto_count_dist.get(v, 0) + 1
        for d in parse_multi(m['q8_4']):
            difficulty_top[d] = difficulty_top.get(d, 0) + 1

    n_regular  = sum(1 for m in members if any(k in (m['q1_3'] or '') for k in ['자주', '거의 항상', '50', '75']))
    n_auto     = sum(1 for m in members if any(k in (m['q5_1'] or '') for k in ['3–5', '3-5', '6개']))
    n_nonshare = sum(1 for m in members if any('공유한 적 없음' in c for c in parse_multi(m['q6_1'])))
    n_time     = sum(1 for m in members if any(k in (m['q5_2'] or '') for k in ['3–6', '3-6', '6–10', '6-10', '10시간']))

    aggs = {
        'tool_usage': tool_usage,
        'freq_dist': freq_dist,
        'share_dist': share_dist,
        'time_saved_dist': time_saved_dist,
        'auto_count_dist': auto_count_dist,
        'difficulty_top': difficulty_top,
        'total_respondents': n_total,
        'kpi': {
            'respondents': n_total, 'regular_users': n_regular,
            'automation_owners': n_auto, 'non_sharers': n_nonshare, 'time_savers': n_time,
        },
    }

    print(f'  정기:{n_regular} 자동화3+:{n_auto} 미공유:{n_nonshare} 시간3h+:{n_time}')

    sum_rows = []
    for i, m in enumerate(members):
        s = sigs[i]
        def cs(sg): return f'<td class="signal-cell {SIG_CLASS[sg[0]]}">{SIG_EMOJI[sg[0]]} {esc(sg[1])}</td>'
        sum_rows.append(
            f'<tr onclick="showMember({i})">\n'
            f'  <td><strong>{esc(m["name"])}</strong></td><td>{esc(m["role"])}</td>\n'
            f'  {cs(s["tool"])}{cs(s["freq"])}{cs(s["case"])}{cs(s["review"])}{cs(s["impact"])}{cs(s["share"])}\n'
            f'  <td class="score-cell score-empty" id="row-total-{i}">—</td>\n'
            f'  <td class="score-cell score-empty" id="row-grade-{i}">—</td>\n'
            f'</tr>'
        )

    tabs = [
        f'<button class="{"active" if i == 0 else ""}" onclick="showMember({i})" id="tab-{i}">{esc(m["name"])}</button>'
        for i, m in enumerate(members)
    ]

    details = []
    for i, m in enumerate(members):
        s = sigs[i]
        cg = sum(1 for k in s.values() if k[0] == 'green')
        cy = sum(1 for k in s.values() if k[0] == 'yellow')
        cr = sum(1 for k in s.values() if k[0] == 'red')
        ax1, ax1s = s['tool'],   s['freq']
        ax2, ax2s = s['case'],   s['review']
        ax3       = s['impact']
        ax4       = s['share']

        # a1 — v7 순서: 1-1(s11) / 1-2(s12) / 1-3(s14, q2.3) / 1-4(s13, q2.2)
        a1 = (
            f'<div class="axis-block {ax1[0]}">'
            f'<div class="axis-header"><h4>01 AI 툴 활용 <span class="axis-weight">(가중치 25%)</span></h4>'
            f'<span class="axis-avg empty" id="axis-avg-{i}-a1">축 평균 <strong>—</strong></span></div>'
            f'<div class="signal-line">{signal_line(ax1, ax1s)}</div>'
            f'<div class="synth-box"><strong>정성 코멘트</strong> — 도구 다양성·정착 깊이·새 도구 시도의 균형을 보세요. 한 도구만 강하게 쓰는 경우 1-2가, 정착이 약한 경우 1-1·1-3이 낮을 수 있습니다.</div>'
            f'<div class="subitem"><div class="subitem-label">1-1. 업무 내 AI 비중 (정기 사용 강도)</div>'
            f'<div class="evidence-list">{evid("q1.3 · AI가 함께 작동하는 업무 비중", m["q1_3"])}</div>{score_row(i, "s11")}</div>'
            f'<div class="subitem"><div class="subitem-label">1-2. 정기 사용 도구 다양성</div>'
            f'<div class="evidence-list">{evid("q2.1 · 정기 사용 도구 (주 2회 이상)", m["q2_1"])}</div>{score_row(i, "s12")}</div>'
            f'<div class="subitem"><div class="subitem-label">1-3. 새 AI 툴 자발 학습·적용</div>'
            f'<div class="evidence-list">{evid("q2.3 · 최근 3개월 신규 도입 도구", m["q2_3"])}</div>{score_row(i, "s14")}</div>'
            f'<div class="subitem"><div class="subitem-label">1-4. AI 사용 목적·방향 자기 설정</div>'
            f'<div class="evidence-list">{evid("q2.2 · 가장 손에 익은 도구·계기", m["q2_2"])}</div>{score_row(i, "s13")}</div>'
            f'</div>'
        )

        # a2 — v7: ⭐ 헤더카드 + s21(q4.1), s22(q4.2)+hint, s23(q4.3)+hint
        header_card_evidence = (
            evid("q3.1 · 잘 풀린 사례 1순위", m["q3_1"])
            + evid("q3.2 · 1순위 산출물 링크", m["q3_2"])
            + evid("q3.3 · 2순위 사례", m["q3_3"])
            + evid("q3.4 · 2순위 산출물 링크", m["q3_4"])
            + evid("q3.5 · 3순위 사례", m["q3_5"])
            + evid("q3.6 · 3순위 산출물 링크", m["q3_6"])
        )
        a2 = (
            f'<div class="axis-block {ax2[0]}">'
            f'<div class="axis-header"><h4>02 결과물 완성도 <span class="axis-weight">(가중치 30%)</span></h4>'
            f'<span class="axis-avg empty" id="axis-avg-{i}-a2">축 평균 <strong>—</strong></span></div>'
            f'<div class="signal-line">{signal_line(ax2, ax2s)}</div>'
            f'<div class="synth-box"><strong>정성 코멘트</strong> — 사례의 구체성·검증 흐름의 체계성이 결과물 완성도의 핵심입니다. 자기 템플릿(2-4) 보유는 강한 신호입니다.</div>'
            f'<div class="subitem axis-header-card" style="background:#fff8e7; border-left:4px solid #f5a623; padding:16px;">'
            f'<div class="subitem-label" style="font-size:1.05em; color:#8a5a00;">⭐ 2축 평가 — 대표 활용 사례 (구체성·다양성)</div>'
            f'<div style="margin:8px 0 12px 0; color:#555; font-size:0.92em; line-height:1.6;">아래 응답을 보고 <strong>2-1, 2-2, 2-3 점수를 매기세요</strong>. 이 카드 자체는 점수 입력 칸이 없습니다 — 사례의 구체성·다양성을 종합 판단의 근거로 사용합니다.</div>'
            f'<div class="evidence-list">{header_card_evidence}</div></div>'
            f'<div class="subitem"><div class="subitem-label">2-1. AI 결과물 오류·편향을 스스로 발견·수정·검증한다</div>'
            f'<div class="evidence-list">{evid("q4.1 · AI 결과물 검토·검증 흐름", m["q4_1"])}</div>{score_row(i, "s21")}</div>'
            f'<div class="subitem"><div class="subitem-label">2-2. AI 활용 산출물의 품질이 기존 대비 향상되었다</div>'
            f'{hint_box(HINT_S22)}'
            f'<div class="evidence-list">{evid("q4.2 · 오류·편향 수정 최근 사례", m["q4_2"])}</div>{score_row(i, "s22")}</div>'
            f'<div class="subitem"><div class="subitem-label">2-3. AI 결과물을 업무 맥락에 맞게 생성할 수 있다</div>'
            f'{hint_box(HINT_S23)}'
            f'<div class="evidence-list">{evid("q4.3 · 자기 체크리스트·템플릿", m["q4_3"])}</div>{score_row(i, "s23")}</div>'
            f'</div>'
        )

        # a3 — v7: s31에 hint 추가
        a3 = (
            f'<div class="axis-block {ax3[0]}">'
            f'<div class="axis-header"><h4>03 속도·임팩트 <span class="axis-weight">(가중치 30%)</span></h4>'
            f'<span class="axis-avg empty" id="axis-avg-{i}-a3">축 평균 <strong>—</strong></span></div>'
            f'<div class="signal-line">{signal_line(ax3)}</div>'
            f'<div class="synth-box"><strong>정성 코멘트</strong> — 시간 체감과 자동화 수는 자기 보고이므로 사례(3-3·3-4)와 함께 봐 주세요. &quot;AI 없으면 못 했을&quot; 사례가 가장 강한 신호입니다.</div>'
            f'<div class="subitem"><div class="subitem-label">3-1. 자동화·시간 단축 반복 업무 수</div>'
            f'{hint_box(HINT_S31)}'
            f'<div class="evidence-list">{evid("q5.1 · 자동화·시간 단축 반복 업무 수", m["q5_1"])}</div>{score_row(i, "s31")}</div>'
            f'<div class="subitem"><div class="subitem-label">3-2. 한 주 절약 시간 (체감)</div>'
            f'<div class="evidence-list">{evid("q5.2 · 한 주 절약 시간 (체감)", m["q5_2"])}</div>{score_row(i, "s32")}</div>'
            f'<div class="subitem"><div class="subitem-label">3-3. AI 덕분에 새로 시도한 업무</div>'
            f'<div class="evidence-list">{evid("q5.3 · 새로 시도한 업무", m["q5_3"])}</div>{score_row(i, "s33")}</div>'
            f'<div class="subitem"><div class="subitem-label">3-4. &quot;AI 없으면 못 했을&quot; 사례</div>'
            f'<div class="evidence-list">{evid("q5.4 · &quot;AI 없으면 못 했을&quot; 사례", m["q5_4"])}{evid("q5.5 · 위 사례의 산출물 링크", m["q5_5"])}</div>{score_row(i, "s34")}</div>'
            f'</div>'
        )

        # a4 — v7: s44 신규 (q-evidence 없음, 리더 종합 판단)
        a4 = (
            f'<div class="axis-block {ax4[0]}">'
            f'<div class="axis-header"><h4>04 지식 공유 <span class="axis-weight">(가중치 15%)</span></h4>'
            f'<span class="axis-avg empty" id="axis-avg-{i}-a4">축 평균 <strong>—</strong></span></div>'
            f'<div class="signal-line">{signal_line(ax4)}</div>'
            f'<div class="synth-box"><strong>정성 코멘트</strong> — 단순 공유(슬랙)와 임팩트 있는 공유(동료 변화)는 점수 차이가 큽니다. 4-2 동료 변화 사례가 빈약하면 4-3 자료 흔적도 보세요.</div>'
            f'<div class="subitem"><div class="subitem-label">4-1. 공유 채널 다양성</div>'
            f'<div class="evidence-list">{evid("q6.1 · 공유 방식 (복수 응답)", m["q6_1"])}</div>{score_row(i, "s41")}</div>'
            f'<div class="subitem"><div class="subitem-label">4-2. 동료 업무 방식 변화 사례</div>'
            f'<div class="evidence-list">{evid("q6.2 · 동료의 업무 방식이 바뀐 사례", m["q6_2"])}</div>{score_row(i, "s42")}</div>'
            f'<div class="subitem"><div class="subitem-label">4-3. 공유 자료 흔적·링크</div>'
            f'<div class="evidence-list">{evid("q6.3 · 공유 자료 대표 링크", m["q6_3"])}</div>{score_row(i, "s43")}</div>'
            f'<div class="subitem"><div class="subitem-label">4-4. AI 활용 관련 팀 내 문화 형성에 기여하고 있다</div>'
            f'{hint_box(HINT_S44)}'
            f'<div class="evidence-list"><div class="evidence-item" style="color:#888; font-style:italic;">4-1, 4-2, 4-3 항목과 멤버 공유 활동을 종합 판단</div></div>{score_row(i, "s44")}</div>'
            f'</div>'
        )

        aux = (
            f'<div class="axis-block">'
            f'<h4>📌 보조 정보 — Red-line · 회고 (점수 영향 없음)</h4>'
            f'{evid("q8.3 · 지금 가장 답답한 지점", m["q8_3"] or "—")}'
            f'{evid("q8.4 · 가장 큰 어려움", m["q8_4"] or "—")}'
            f'{evid("q8.5 · 필요한 지원", m["q8_5"] or "—")}'
            f'{evid("q8.6 · 가장 자신 있는 AI 활용 영역", m["q8_6"] or "—")}'
            f'{evid("q8.7 · 다음 3개월 시도하고 싶은 것", m["q8_7"] or "—")}'
            f'</div>'
        )

        active = 'active' if i == 0 else ''
        role_line = f'{esc(m["role"])} · {esc(m["team"])} · 리더: {esc(m["leader"])}'

        details.append(
            f'<div class="member-detail {active}" id="mem-{i}">'
            f'<div class="member-header">'
            f'<div class="name-block"><h3>{esc(m["name"])}</h3><div class="role">{role_line}</div></div>'
            f'<div class="signal-strip">'
            f'<span class="sig-chip">🟢 {cg}</span><span class="sig-chip">🟡 {cy}</span><span class="sig-chip">🔴 {cr}</span>'
            f'</div></div>'
            f'<div class="score-summary">'
            f'<div class="total">총점 <strong id="total-{i}">—</strong> / 100</div>'
            f'<div class="actions">'
            f'<span class="saved-badge" id="saved-badge-{i}">✓ 저장됨</span>'
            f'<div class="grade empty" id="grade-{i}">등급 미입력</div>'
            f'<button class="btn-save" onclick="saveMember({i})">💾 저장</button>'
            f'<button class="btn-delete" onclick="deleteMember({i})">🗑 삭제</button>'
            f'</div></div>'
            f'{a1}{a2}{a3}{a4}{aux}'
            f'<div class="memo-block">'
            f'<div class="score-label">📝 리더 메모 (1:1 코칭 포커스·결정 근거 등)</div>'
            f'<textarea class="memo-input" id="memo-{i}" placeholder="자유롭게 입력하세요. 자동 저장됩니다." oninput="setMemo({i}, this.value)"></textarea>'
            f'</div>'
            f'<div class="rubric-aside"><strong>리더 체크 포인트</strong> — 자기 보고는 과대·과소 편향이 모두 가능합니다. 평소 1:1·산출물 리뷰·동료 피드백을 함께 두고 세부 항목을 매겨 주세요. AI는 점수를 제안하지 않습니다.</div>'
            f'</div>'
        )

    members_json = json.dumps(
        [{'name': m['name'], 'role': m['role'], 'team': m['team'], 'leader': m['leader']} for m in members],
        ensure_ascii=False,
    )
    aggs_json = json.dumps(aggs, ensure_ascii=False)
    storage_key = f'{dept}_aiq_v7'
    filename_base = f'{dept}_AIQ_세부평가결과'

    html_out = HTML_TEMPLATE.format(
        DEPT=dept,
        n_total=n_total,
        n_regular=n_regular,
        n_auto=n_auto,
        n_nonshare=n_nonshare,
        n_time=n_time,
        sum_rows='\n'.join(sum_rows),
        tabs='\n      '.join(tabs),
        details='\n'.join(details),
        aggs_json=aggs_json,
        members_json=members_json,
        storage_key=storage_key,
        filename_base=filename_base,
    )

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html_out)
    print(f'✅ 저장: {output_path} ({os.path.getsize(output_path):,} bytes)')


# ---------- HTML 템플릿 ----------
_HTML_RAW = r"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>__DEPT__ · AI-Q 리더 평가 가이드 (v7) (세부 항목별 점수)</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script>
<style>
* { box-sizing: border-box; }
body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', 'Noto Sans KR', sans-serif; margin: 0; padding: 0; background: #f7f8fa; color: #1a1a1a; line-height: 1.6; }
.container { max-width: 1280px; margin: 0 auto; padding: 32px 24px; }
header.page-header { background: linear-gradient(135deg, #1a1a1a 0%, #2a2a2a 100%); color: #fff; padding: 40px 24px; }
header .container { padding-top: 0; padding-bottom: 0; }
header h1 { margin: 0 0 8px; font-size: 28px; font-weight: 700; }
header .sub { color: #ffd400; font-size: 14px; margin: 0 0 16px; }
header .desc { color: #cfcfcf; font-size: 14px; max-width: 800px; line-height: 1.7; }
.warning-banner { background: #fff8dc; border-left: 4px solid #ffd400; padding: 16px 20px; margin: 24px 0; border-radius: 4px; font-size: 14px; }
.warning-banner strong { color: #1a1a1a; }
section.block { background: #fff; border-radius: 12px; padding: 28px; margin-bottom: 24px; box-shadow: 0 1px 3px rgba(0,0,0,0.04); }
section.block h2 { margin: 0 0 8px; font-size: 20px; font-weight: 700; }
section.block .section-sub { color: #666; font-size: 13px; margin-bottom: 20px; }
.kpi-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr)); gap: 12px; margin-bottom: 20px; }
.kpi-card { background: #fafafa; border-radius: 8px; padding: 16px; border-left: 4px solid #ffd400; }
.kpi-card .label { font-size: 11px; color: #666; margin-bottom: 4px; }
.kpi-card .value { font-size: 22px; font-weight: 700; color: #1a1a1a; }
.kpi-card .sub { font-size: 11px; color: #888; margin-top: 4px; }
.chart-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 24px; }
.chart-grid > div { background: #fafafa; border-radius: 8px; padding: 20px; }
.chart-grid h3 { font-size: 14px; margin: 0 0 12px; }
.chart-wrap { position: relative; height: 240px; }
.bar-list { list-style: none; padding: 0; margin: 0; }
.bar-list li { display: flex; align-items: center; gap: 12px; margin-bottom: 8px; font-size: 13px; }
.bar-list .bar-name { flex: 0 0 130px; color: #444; }
.bar-list .bar-bg { flex: 1; height: 18px; background: #ececec; border-radius: 4px; overflow: hidden; }
.bar-list .bar-fill { height: 100%; background: #ffd400; transition: width 0.4s; }
.bar-list .bar-count { flex: 0 0 36px; text-align: right; font-weight: 600; color: #1a1a1a; }
table.summary-table { width: 100%; border-collapse: collapse; font-size: 13px; }
table.summary-table th, table.summary-table td { padding: 10px 12px; text-align: left; border-bottom: 1px solid #eee; }
table.summary-table th { background: #fafafa; font-weight: 600; color: #444; }
table.summary-table tr:hover { background: #fffbe6; cursor: pointer; }
.signal-cell { font-size: 12px; }
.sig-green { color: #1a7a3c; }
.sig-yellow { color: #b58900; }
.sig-red { color: #c0392b; }
.score-cell { font-weight: 700; font-size: 13px; }
.score-empty { color: #ccc; }
.member-tabs { display: flex; flex-wrap: wrap; gap: 6px; margin-bottom: 20px; padding: 16px; background: #fafafa; border-radius: 8px; }
.member-tabs button { padding: 6px 12px; border: 1px solid #ddd; background: #fff; border-radius: 16px; font-size: 12px; cursor: pointer; transition: all 0.15s; }
.member-tabs button:hover { background: #fff8dc; border-color: #ffd400; }
.member-tabs button.active { background: #ffd400; border-color: #ffd400; font-weight: 600; }
.member-tabs button.scored { background: #e8f5e9; border-color: #4caf50; }
.member-tabs button.confirmed { background: #1a7a3c; color: #fff; border-color: #1a7a3c; }
.member-tabs button.active, .member-tabs button.confirmed.active { background: #ffd400; color: #1a1a1a; border-color: #ffd400; }
.member-detail { display: none; }
.member-detail.active { display: block; }
.member-header { display: flex; justify-content: space-between; align-items: flex-start; flex-wrap: wrap; gap: 16px; margin-bottom: 24px; padding-bottom: 16px; border-bottom: 2px solid #ffd400; }
.member-header .name-block h3 { margin: 0 0 4px; font-size: 22px; }
.member-header .name-block .role { color: #666; font-size: 13px; }
.signal-strip { display: flex; gap: 8px; flex-wrap: wrap; }
.signal-strip .sig-chip { padding: 6px 12px; border-radius: 16px; font-size: 12px; background: #fafafa; border: 1px solid #eee; }
.score-summary { background: #1a1a1a; color: #fff; padding: 16px 20px; border-radius: 8px; margin-bottom: 20px; display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 16px; }
.score-summary .total { font-size: 14px; }
.score-summary .total strong { color: #ffd400; font-size: 22px; margin-left: 8px; }
.score-summary .grade { padding: 6px 16px; background: #ffd400; color: #1a1a1a; border-radius: 16px; font-weight: 700; font-size: 18px; }
.score-summary .grade.empty { background: #444; color: #aaa; font-size: 13px; font-weight: normal; }
.score-summary .actions { display: flex; gap: 8px; align-items: center; }
.score-summary .actions button { padding: 7px 14px; border-radius: 4px; border: none; font-weight: 600; cursor: pointer; font-size: 12px; transition: opacity 0.15s; }
.score-summary .actions .btn-save { background: #ffd400; color: #1a1a1a; }
.score-summary .actions .btn-delete { background: #c0392b; color: #fff; }
.score-summary .actions button:hover { opacity: 0.85; }
.score-summary .saved-badge { padding: 4px 10px; background: #1a7a3c; color: #fff; border-radius: 12px; font-size: 11px; font-weight: 600; display: none; }
.score-summary .saved-badge.show { display: inline-block; }
.axis-block { margin-bottom: 28px; padding: 20px; background: #fafafa; border-radius: 8px; border-left: 4px solid #ddd; }
.axis-block.green { border-left-color: #1a7a3c; }
.axis-block.yellow { border-left-color: #b58900; }
.axis-block.red { border-left-color: #c0392b; }
.axis-header { display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 12px; margin-bottom: 8px; }
.axis-header h4 { margin: 0; font-size: 15px; display: flex; align-items: center; gap: 10px; }
.axis-header .axis-weight { color: #888; font-size: 12px; font-weight: normal; }
.axis-avg { padding: 4px 12px; background: #fff; border: 1px solid #ddd; border-radius: 16px; font-size: 12px; font-weight: 600; }
.axis-avg.empty { color: #aaa; font-weight: normal; }
.axis-avg strong { color: #1a1a1a; font-size: 14px; margin-left: 6px; }
.axis-block .signal-line { font-size: 13px; color: #555; margin: 8px 0 16px; }
.subitem { margin-bottom: 16px; padding: 14px 16px; background: #fff; border-radius: 6px; border: 1px solid #eee; }
.subitem .subitem-label { font-size: 13px; font-weight: 700; color: #1a1a1a; margin-bottom: 10px; }
.subitem .evidence-list { margin-bottom: 12px; }
.evidence-item { margin-bottom: 8px; padding: 8px 12px; background: #fafafa; border-radius: 4px; border-left: 2px solid #ddd; }
.evidence-item .q { font-size: 11px; color: #888; margin-bottom: 2px; font-weight: 600; }
.evidence-item .a { font-size: 13px; color: #222; white-space: pre-wrap; word-break: break-word; }
.subitem .score-row { display: flex; align-items: center; gap: 10px; flex-wrap: wrap; padding-top: 8px; border-top: 1px solid #eee; }
.subitem .score-row .score-prompt { font-size: 12px; color: #666; font-weight: 600; }
.score-buttons { display: flex; gap: 4px; align-items: center; }
.score-buttons button { width: 32px; height: 32px; border-radius: 50%; border: 2px solid #ddd; background: #fff; font-weight: 700; cursor: pointer; transition: all 0.15s; font-size: 12px; }
.score-buttons button:hover { border-color: #ffd400; background: #fff8dc; }
.score-buttons button.selected { background: #1a1a1a; color: #ffd400; border-color: #1a1a1a; }
.score-buttons .clear-btn { width: auto; padding: 0 10px; height: 28px; font-size: 11px; font-weight: normal; color: #888; border-radius: 14px; border-width: 1px; }
.synth-box { margin-top: 6px; margin-bottom: 16px; padding: 12px 14px; background: #fff8dc; border-radius: 6px; border-left: 3px solid #ffd400; font-size: 12px; color: #4a3c00; }
.synth-box strong { color: #1a1a1a; }
.memo-input { width: 100%; margin-top: 10px; padding: 8px 12px; border: 1px solid #ddd; border-radius: 6px; font-size: 12px; font-family: inherit; min-height: 60px; resize: vertical; }
.memo-block { margin-top: 16px; padding: 14px 16px; background: #fff; border: 2px dashed #ffd400; border-radius: 6px; }
.memo-block .score-label { font-size: 12px; color: #666; margin-bottom: 8px; font-weight: 600; }
.rubric-aside { background: #f4f4f4; padding: 16px; border-radius: 8px; font-size: 12px; color: #555; margin-top: 24px; line-height: 1.7; }
.rubric-aside strong { color: #1a1a1a; }
.guide-toggle { display: inline-block; margin-bottom: 16px; padding: 8px 16px; background: #1a1a1a; color: #fff; border: none; border-radius: 4px; font-size: 12px; cursor: pointer; }
.guide-content { display: none; background: #fafafa; padding: 20px; border-radius: 8px; margin-bottom: 20px; font-size: 13px; line-height: 1.8; }
.guide-content.show { display: block; }
.guide-content table { width: 100%; margin: 12px 0; }
.guide-content th, .guide-content td { padding: 8px; border: 1px solid #ddd; text-align: left; }
.action-bar { position: sticky; top: 0; z-index: 100; background: #1a1a1a; color: #fff; padding: 12px 24px; display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 12px; }
.action-bar .progress { font-size: 13px; color: #cfcfcf; }
.action-bar .progress strong { color: #ffd400; }
.action-bar button { padding: 8px 16px; background: #ffd400; color: #1a1a1a; border: none; border-radius: 4px; font-weight: 600; cursor: pointer; font-size: 13px; }
.action-bar button.secondary { background: #444; color: #fff; }
.action-bar button:hover { opacity: 0.9; }
.toast { position: fixed; bottom: 24px; left: 50%; transform: translateX(-50%); background: #1a1a1a; color: #fff; padding: 12px 24px; border-radius: 24px; font-size: 13px; opacity: 0; transition: opacity 0.3s; pointer-events: none; z-index: 1000; }
.toast.show { opacity: 1; }
@media (max-width: 900px) { .chart-grid { grid-template-columns: 1fr; } }
</style>
</head>
<body>
<header class="page-header">
  <div class="container">
    <h1>__DEPT__ · AI-Q 리더 평가 가이드</h1>
    <p class="sub">세부 항목별 점수 입력 (4축 × 3~4 sub-item)</p>
    <p class="desc">자가진단 폼 설계에 정의된 세부 항목 단위로 1~5점을 매기시면 축 평균이 자동 계산되고, 4축 가중 평균으로 총점이 산출됩니다. AI는 응답 패턴(🟢🟡🔴)만 정성적으로 보여드리며 점수는 제안하지 않습니다.</p>
  </div>
</header>

<div class="action-bar">
  <div class="progress">평가 진행: <strong id="progress-text">0 / __N_TOTAL__</strong></div>
  <div>
    <button onclick="exportExcel()">📥 엑셀(XLSX) 내보내기</button>
    <button class="secondary" onclick="exportCSV()">📥 CSV 내보내기</button>
    <button class="secondary" onclick="resetAll()">↺ 전체 초기화</button>
  </div>
</div>

<div class="container">
  <div class="warning-banner">
    ⚠️ <strong>AI는 점수를 제안하지 않습니다.</strong> 세부 항목별 점수는 평소 관찰·산출물·자가진단 응답을 종합해 리더께서 직접 매겨 주세요.
  </div>

  <button class="guide-toggle" onclick="document.getElementById('guide').classList.toggle('show')">▾ 평가 가이드 펼치기</button>
  <div class="guide-content" id="guide">
    <h3 style="margin-top:0;">AI-Q 4축 · 가중치 · 세부 항목</h3>
    <table>
      <tr><th>축</th><th>가중치</th><th>세부 항목</th></tr>
      <tr><td>01 AI 툴 활용</td><td>25%</td><td>1-1 비중 / 1-2 도구 다양성 / 1-3 새 도구 학습 / 1-4 목적·방향 설정</td></tr>
      <tr><td>02 결과물 완성도</td><td>30%</td><td>⭐사례 헤더카드 / 2-1 오류·편향 검증 / 2-2 품질 향상 / 2-3 맥락 반영</td></tr>
      <tr><td>03 속도·임팩트</td><td>30%</td><td>3-1 자동화 수 / 3-2 절약 시간 / 3-3 신규 시도 / 3-4 AI 없으면 못 했을 사례</td></tr>
      <tr><td>04 지식 공유</td><td>15%</td><td>4-1 채널 다양성 / 4-2 동료 변화 / 4-3 자료 흔적 / 4-4 문화 형성 기여</td></tr>
    </table>
    <p style="font-size:12px; color:#666;">
      축 점수 = 세부 항목 평균<br>
      총점 = (축1·25% + 축2·30% + 축3·30% + 축4·15%) × 20<br>
      등급: S(85+) / A(70~85) / B(55~70) / C(&lt;55)
    </p>
    <h3>신호 해석 (정성·점수 무관)</h3>
    <p>
      <span class="sig-green"><strong>🟢 정착·체화</strong></span> — 응답에 구체적 사례·다양성·일상 흐름이 보임<br>
      <span class="sig-yellow"><strong>🟡 진행</strong></span> — 신호는 있으나 사례 수·구체성·정착도가 한정적<br>
      <span class="sig-red"><strong>🔴 시작·간헐</strong></span> — 사례·정착·공유 흔적 부족
    </p>
  </div>

  <section class="block">
    <h2>📊 팀 대시보드</h2>
    <p class="section-sub">__N_TOTAL__명 응답 집계 — __DEPT__ 전체 분포</p>
    <div class="kpi-grid">
      <div class="kpi-card"><div class="label">응답자</div><div class="value">__N_TOTAL__명</div><div class="sub">팀 전체</div></div>
      <div class="kpi-card"><div class="label">AI 정기 사용</div><div class="value">__N_REGULAR__명</div><div class="sub">업무 50% 이상</div></div>
      <div class="kpi-card"><div class="label">자동화 보유자</div><div class="value">__N_AUTO__명</div><div class="sub">3건 이상</div></div>
      <div class="kpi-card"><div class="label">미공유</div><div class="value">__N_NONSHARE__명</div><div class="sub">공유 활동 없음</div></div>
      <div class="kpi-card"><div class="label">시간 절약 ≥ 3h/주</div><div class="value">__N_TIME__명</div><div class="sub">체감 응답</div></div>
    </div>
    <div class="chart-grid">
      <div><h3>도구 사용 빈도 (정기 사용·복수 응답)</h3><ul class="bar-list" id="tools-bar"></ul></div>
      <div><h3>AI 업무 비중 분포</h3><div class="chart-wrap"><canvas id="freqChart"></canvas></div></div>
      <div><h3>주 절약 시간 분포</h3><div class="chart-wrap"><canvas id="timeChart"></canvas></div></div>
      <div><h3>공유 활동 (복수 응답)</h3><ul class="bar-list" id="share-bar"></ul></div>
      <div><h3>가장 큰 어려움 (복수 응답)</h3><ul class="bar-list" id="diff-bar"></ul></div>
      <div><h3>자동화 보유 건수</h3><div class="chart-wrap"><canvas id="autoChart"></canvas></div></div>
    </div>
  </section>

  <section class="block">
    <h2>👥 멤버별 신호 + 점수 요약</h2>
    <p class="section-sub">행을 클릭하면 해당 멤버 상세로 이동합니다. 신호는 정성 패턴, 점수는 리더 입력값입니다.</p>
    <table class="summary-table">
      <thead>
        <tr><th>이름</th><th>직책</th><th>01 툴</th><th>01 빈도</th><th>02 사례</th><th>02 검토</th><th>03 임팩트</th><th>04 공유</th><th>총점</th><th>등급</th></tr>
      </thead>
      <tbody id="summary-body">
__SUM_ROWS__
      </tbody>
    </table>
  </section>

  <section class="block">
    <h2>🔍 멤버별 응답 + 세부 항목별 점수 입력</h2>
    <p class="section-sub">각 세부 항목 옆 1~5 버튼으로 점수를 입력하면 축 평균과 총점이 자동 계산됩니다.</p>
    <div class="member-tabs" id="member-tabs">
      __TABS__
    </div>
__DETAILS__
  </section>

  <div class="rubric-aside">
    <p>작성: NGA AI-Q v7 · __DEPT__ 부서 평가지 (sub-item 점수 + 멤버별 저장·삭제)</p>
  </div>
</div>

<script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
<script>
const aggs = __AGGS_JSON__;
const members = __MEMBERS_JSON__;
const SUB_KEYS = ["s11","s12","s13","s14","s21","s22","s23","s31","s32","s33","s34","s41","s42","s43","s44"];
const AXIS_SUB = {"a1":["s11","s12","s13","s14"],"a2":["s21","s22","s23"],"a3":["s31","s32","s33","s34"],"a4":["s41","s42","s43","s44"]};
const SUB_LABELS = {"s11":"1-1. 업무 내 AI 비중 (정기 사용 강도)","s12":"1-2. 정기 사용 도구 다양성","s13":"1-3. 새 AI 툴 자발 학습·적용","s14":"1-4. AI 사용 목적·방향 자기 설정","s21":"2-1. AI 결과물 오류·편향 발견·수정·검증","s22":"2-2. 산출물 품질 향상","s23":"2-3. 업무 맥락 반영 생성","s31":"3-1. 자동화·시간 단축 반복 업무 수","s32":"3-2. 한 주 절약 시간 (체감)","s33":"3-3. AI 덕분에 새로 시도한 업무","s34":"3-4. \"AI 없으면 못 했을\" 사례","s41":"4-1. 공유 채널 다양성","s42":"4-2. 동료 업무 방식 변화 사례","s43":"4-3. 공유 자료 흔적·링크","s44":"4-4. 팀 내 AI 문화 형성 기여"};
const WEIGHTS = { a1: 0.25, a2: 0.30, a3: 0.30, a4: 0.15 };
const STORAGE_KEY = '__STORAGE_KEY__';
const FILENAME = '__FILENAME_BASE__';
let state = loadState();

function loadState() {
  try {
    const raw = localStorage.getItem(STORAGE_KEY);
    if (raw) { const s = JSON.parse(raw); if (!s.saved) s.saved = {}; return s; }
  } catch(e) {}
  const init = { scores: {}, memos: {}, saved: {} };
  members.forEach((_, i) => { init.scores[i] = {}; SUB_KEYS.forEach(k => init.scores[i][k] = null); init.memos[i] = ''; });
  return init;
}
function saveState() { localStorage.setItem(STORAGE_KEY, JSON.stringify(state)); }
function toast(msg) {
  let el = document.getElementById('toast');
  if (!el) { el = document.createElement('div'); el.id = 'toast'; el.className = 'toast'; document.body.appendChild(el); }
  el.textContent = msg; el.classList.add('show');
  clearTimeout(window.__toastTimer);
  window.__toastTimer = setTimeout(() => el.classList.remove('show'), 1800);
}
function setScore(memberIdx, subKey, value) {
  if (!state.scores[memberIdx]) state.scores[memberIdx] = {};
  state.scores[memberIdx][subKey] = value;
  saveState();
  renderSubButtons(memberIdx, subKey); renderAxisAvg(memberIdx);
  renderSummary(memberIdx); renderSummaryRow(memberIdx);
  renderTabClass(memberIdx); renderProgress();
}
function setMemo(memberIdx, value) { state.memos[memberIdx] = value; saveState(); }
function avgFilled(memberIdx, subKeys) {
  const s = state.scores[memberIdx] || {};
  const vals = subKeys.map(k => s[k]).filter(v => v !== null && v !== undefined);
  if (vals.length === 0) return null;
  return vals.reduce((a,b)=>a+b,0) / vals.length;
}
function axisAvg(memberIdx, axisKey) { return avgFilled(memberIdx, AXIS_SUB[axisKey]); }
function allSubsFilled(memberIdx) {
  const s = state.scores[memberIdx] || {};
  return SUB_KEYS.every(k => s[k] !== null && s[k] !== undefined);
}
function calcTotal(memberIdx) {
  const a1 = axisAvg(memberIdx, 'a1'), a2 = axisAvg(memberIdx, 'a2'), a3 = axisAvg(memberIdx, 'a3'), a4 = axisAvg(memberIdx, 'a4');
  if ([a1,a2,a3,a4].some(v => v === null)) return { total: null, grade: null };
  const total = (a1 * WEIGHTS.a1 + a2 * WEIGHTS.a2 + a3 * WEIGHTS.a3 + a4 * WEIGHTS.a4) * 20;
  let grade = 'C';
  if (total >= 85) grade = 'S'; else if (total >= 70) grade = 'A'; else if (total >= 55) grade = 'B';
  return { total: Math.round(total * 10) / 10, grade };
}
function renderSubButtons(memberIdx, subKey) {
  const wrap = document.querySelector(`.score-buttons[data-member="${memberIdx}"][data-sub="${subKey}"]`);
  if (!wrap) return;
  const v = (state.scores[memberIdx] || {})[subKey];
  wrap.querySelectorAll('button').forEach((btn, idx) => {
    if (idx < 5) btn.classList.toggle('selected', v === idx + 1);
  });
}
function renderAxisAvg(memberIdx) {
  ['a1','a2','a3','a4'].forEach(axisKey => {
    const el = document.getElementById(`axis-avg-${memberIdx}-${axisKey}`);
    if (!el) return;
    const v = axisAvg(memberIdx, axisKey);
    if (v === null) { el.classList.add('empty'); el.innerHTML = '축 평균 <strong>—</strong>'; }
    else { el.classList.remove('empty'); el.innerHTML = `축 평균 <strong>${v.toFixed(2)}</strong> / 5`; }
  });
}
function renderSummary(memberIdx) {
  const { total, grade } = calcTotal(memberIdx);
  const totalEl = document.getElementById('total-' + memberIdx);
  const gradeEl = document.getElementById('grade-' + memberIdx);
  if (total === null) { totalEl.textContent = '—'; gradeEl.textContent = '등급 미입력'; gradeEl.classList.add('empty'); }
  else { totalEl.textContent = total; gradeEl.textContent = grade + ' 등급'; gradeEl.classList.remove('empty'); }
}
function renderSummaryRow(memberIdx) {
  const { total, grade } = calcTotal(memberIdx);
  const tEl = document.getElementById('row-total-' + memberIdx), gEl = document.getElementById('row-grade-' + memberIdx);
  if (total === null) { tEl.textContent = '—'; tEl.classList.add('score-empty'); gEl.textContent = '—'; gEl.classList.add('score-empty'); }
  else { tEl.textContent = total; tEl.classList.remove('score-empty'); gEl.textContent = grade; gEl.classList.remove('score-empty'); }
}
function renderTabClass(memberIdx) {
  const tab = document.getElementById('tab-' + memberIdx);
  const filled = allSubsFilled(memberIdx), confirmed = state.saved && state.saved[memberIdx];
  tab.classList.toggle('confirmed', !!confirmed);
  tab.classList.toggle('scored', filled && !confirmed);
}
function renderSavedBadge(memberIdx) {
  const el = document.getElementById('saved-badge-' + memberIdx); if (!el) return;
  if (state.saved && state.saved[memberIdx]) el.classList.add('show'); else el.classList.remove('show');
}
function renderMemo(memberIdx) { const el = document.getElementById('memo-' + memberIdx); if (el) el.value = state.memos[memberIdx] || ''; }
function renderProgress() {
  let done = 0, saved = 0;
  members.forEach((_, i) => { if (allSubsFilled(i)) done++; if (state.saved && state.saved[i]) saved++; });
  document.getElementById('progress-text').textContent = `입력완료 ${done} / 저장 ${saved} / ${members.length}`;
}
function saveMember(idx) {
  state.saved[idx] = new Date().toISOString(); saveState();
  renderSavedBadge(idx); renderTabClass(idx); renderProgress();
  toast(`✓ ${members[idx].name} 평가 저장됨`);
}
function deleteMember(idx) {
  if (!confirm(`${members[idx].name} 멤버의 점수·메모·저장 기록을 모두 삭제하시겠습니까?`)) return;
  state.scores[idx] = {}; SUB_KEYS.forEach(k => state.scores[idx][k] = null);
  state.memos[idx] = ''; delete state.saved[idx]; saveState();
  SUB_KEYS.forEach(k => renderSubButtons(idx, k));
  renderAxisAvg(idx); renderSummary(idx); renderSummaryRow(idx);
  renderTabClass(idx); renderSavedBadge(idx); renderMemo(idx); renderProgress();
  toast(`🗑 ${members[idx].name} 평가 삭제됨`);
}
function showMember(idx) {
  document.querySelectorAll('.member-detail').forEach(d => d.classList.remove('active'));
  document.querySelectorAll('.member-tabs button').forEach(b => b.classList.remove('active'));
  document.getElementById('mem-' + idx).classList.add('active');
  const tab = document.getElementById('tab-' + idx); tab.classList.add('active');
  document.getElementById('mem-' + idx).scrollIntoView({ behavior: 'smooth', block: 'start' });
  tab.scrollIntoView({ behavior: 'smooth', block: 'nearest', inline: 'center' });
}
function exportCSV() {
  const headers = ['이름','직책','팀','리더'];
  SUB_KEYS.forEach(k => headers.push(SUB_LABELS[k]));
  ['a1','a2','a3','a4'].forEach(a => headers.push(`${a.toUpperCase()} 축평균`));
  headers.push('총점','등급','메모');
  let csv = '﻿' + headers.map(h=>`"${h}"`).join(',') + '\n';
  members.forEach((m, i) => {
    const s = state.scores[i] || {};
    const { total, grade } = calcTotal(i);
    const a1 = axisAvg(i,'a1'), a2 = axisAvg(i,'a2'), a3 = axisAvg(i,'a3'), a4 = axisAvg(i,'a4');
    const row = [m.name, m.role, m.team, m.leader || ''];
    SUB_KEYS.forEach(k => row.push(s[k] !== null && s[k] !== undefined ? s[k] : ''));
    row.push(a1 !== null ? a1.toFixed(2) : '', a2 !== null ? a2.toFixed(2) : '', a3 !== null ? a3.toFixed(2) : '', a4 !== null ? a4.toFixed(2) : '');
    row.push(total || '', grade || '', (state.memos[i] || '').replace(/\n/g, ' '));
    csv += row.map(v => `"${String(v).replace(/"/g,'""')}"`).join(',') + '\n';
  });
  const blob = new Blob([csv], { type: 'text/csv;charset=utf-8' });
  const url = URL.createObjectURL(blob);
  const a = document.createElement('a'); a.href = url; a.download = FILENAME + '.csv'; a.click();
  URL.revokeObjectURL(url);
}
function exportExcel() {
  const headers = ['이름','직책','팀','리더'];
  SUB_KEYS.forEach(k => headers.push(SUB_LABELS[k]));
  ['a1','a2','a3','a4'].forEach(a => headers.push(`${a.toUpperCase()} 축평균`));
  headers.push('총점','등급','리더 메모');
  const rows = [headers];
  members.forEach((m, i) => {
    const s = state.scores[i] || {};
    const { total, grade } = calcTotal(i);
    const a1 = axisAvg(i,'a1'), a2 = axisAvg(i,'a2'), a3 = axisAvg(i,'a3'), a4 = axisAvg(i,'a4');
    const row = [m.name, m.role, m.team, m.leader || ''];
    SUB_KEYS.forEach(k => row.push(s[k] !== null && s[k] !== undefined ? s[k] : ''));
    row.push(a1 !== null ? +a1.toFixed(2) : '', a2 !== null ? +a2.toFixed(2) : '', a3 !== null ? +a3.toFixed(2) : '', a4 !== null ? +a4.toFixed(2) : '');
    row.push(total || '', grade || '', state.memos[i] || '');
    rows.push(row);
  });
  const ws = XLSX.utils.aoa_to_sheet(rows);
  const cols = [{wch:14},{wch:10},{wch:14},{wch:10}];
  for (let i = 0; i < SUB_KEYS.length; i++) cols.push({wch:8});
  for (let i = 0; i < 4; i++) cols.push({wch:10});
  cols.push({wch:8},{wch:6},{wch:40});
  ws['!cols'] = cols;
  const wb = XLSX.utils.book_new();
  XLSX.utils.book_append_sheet(wb, ws, 'AI-Q 세부평가');
  XLSX.writeFile(wb, FILENAME + '.xlsx');
}
function resetAll() {
  if (!confirm('모든 점수·메모·저장 기록을 초기화하시겠습니까? 되돌릴 수 없습니다.')) return;
  localStorage.removeItem(STORAGE_KEY); state = loadState();
  members.forEach((_, i) => {
    SUB_KEYS.forEach(k => renderSubButtons(i, k));
    renderAxisAvg(i); renderSummary(i); renderSummaryRow(i);
    renderTabClass(i); renderSavedBadge(i); renderMemo(i);
  });
  renderProgress();
}
function renderBar(elemId, data, maxItems) {
  const el = document.getElementById(elemId); if (!el) return;
  const items = Object.entries(data).sort((a,b)=>b[1]-a[1]).slice(0, maxItems || 8);
  if (items.length === 0) { el.innerHTML = '<li style="color:#aaa">응답 없음</li>'; return; }
  const max = Math.max(...items.map(x=>x[1]));
  el.innerHTML = items.map(([name, count]) => {
    const w = (count/max*100).toFixed(0);
    let displayName = name.length > 22 ? name.slice(0,20)+'…' : name;
    return `<li><span class="bar-name" title="${name.replace(/"/g,'&quot;')}">${displayName}</span><div class="bar-bg"><div class="bar-fill" style="width:${w}%"></div></div><span class="bar-count">${count}</span></li>`;
  }).join('');
}
renderBar('tools-bar', aggs.tool_usage, 8);
renderBar('share-bar', aggs.share_dist, 8);
renderBar('diff-bar', aggs.difficulty_top, 8);
const chartOpts = { responsive: true, maintainAspectRatio: false, plugins: { legend: { position: 'right', labels: { boxWidth: 12, font: { size: 11 } } } } };
const colors = ['#ffd400', '#1a7a3c', '#b58900', '#c0392b', '#888', '#ccc'];
function donut(canvasId, data, order) {
  const el = document.getElementById(canvasId); if (!el) return;
  const labels = order.filter(k => data[k] !== undefined);
  if (labels.length === 0) return;
  const values = labels.map(k => data[k]);
  new Chart(el, { type: 'doughnut', data: { labels, datasets: [{ data: values, backgroundColor: colors }] }, options: chartOpts });
}
donut('freqChart', aggs.freq_dist, ['거의 항상 (75% 이상)', '자주 (50–75%)', '꾸준히 (25–50%)', '가끔 (10–25%)', '거의 사용하지 않음 (10% 미만)', '거의 없음 (10% 미만)']);
donut('timeChart', aggs.time_saved_dist, ['10시간 이상', '6–10시간', '3–6시간', '1–3시간', '1시간 미만', '거의 없음']);
donut('autoChart', aggs.auto_count_dist, ['6개 이상', '3–5개', '1–2개', '0개']);
members.forEach((_, i) => {
  SUB_KEYS.forEach(k => renderSubButtons(i, k));
  renderAxisAvg(i); renderSummary(i); renderSummaryRow(i);
  renderTabClass(i); renderSavedBadge(i); renderMemo(i);
});
renderProgress();
</script>
</body>
</html>
"""


class _Template:
    def __init__(self, raw):
        self.raw = raw

    def format(self, **kw):
        out = self.raw
        replacements = {
            '__DEPT__':         kw['DEPT'],
            '__N_TOTAL__':      str(kw['n_total']),
            '__N_REGULAR__':    str(kw['n_regular']),
            '__N_AUTO__':       str(kw['n_auto']),
            '__N_NONSHARE__':   str(kw['n_nonshare']),
            '__N_TIME__':       str(kw['n_time']),
            '__SUM_ROWS__':     kw['sum_rows'],
            '__TABS__':         kw['tabs'],
            '__DETAILS__':      kw['details'],
            '__AGGS_JSON__':    kw['aggs_json'],
            '__MEMBERS_JSON__': kw['members_json'],
            '__STORAGE_KEY__':  kw['storage_key'],
            '__FILENAME_BASE__': kw['filename_base'],
        }
        for k, v in replacements.items():
            out = out.replace(k, v)
        return out


HTML_TEMPLATE = _Template(_HTML_RAW)


def main():
    ap = argparse.ArgumentParser(description='AI-Q 리더 평가 가이드 v7 HTML 빌더')
    ap.add_argument('--input', required=True)
    ap.add_argument('--department', required=True)
    ap.add_argument('--output', default=None)
    args = ap.parse_args()

    out = args.output or os.path.join(
        os.path.dirname(os.path.abspath(args.input)),
        f'{args.department}_AIQ_리더평가가이드.html'
    )
    build(args.input, args.department, out)


if __name__ == '__main__':
    main()
